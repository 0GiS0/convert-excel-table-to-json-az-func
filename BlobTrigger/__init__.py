import logging
import azure.functions as func

from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import json


def main(myblob: func.InputStream):
    logging.info(f"Python blob trigger function processed blob \n"
                 f"Name: {myblob.name}\n"
                 f"Blob Size: {myblob.length} bytes")

    # Load the workbook
    workbook = load_workbook(filename=BytesIO(myblob.read()))
    logging.info(f"Loaded workbook: {workbook}")

    # Get the active sheet
    worksheet = workbook.active

    excel_data = []

    # Iterate through the rows and columns of the worksheet
    last_column = len(list(worksheet.columns))
    last_row = len(list(worksheet.rows))

    # Iterate row
    for row in range(1, last_row + 1):
        row_data = {}
        # Iterate column
        for column in range(1, last_column + 1):
            # Get the value of the cell
            column_letter = get_column_letter(column)
            if row > 1:
                row_data[worksheet[column_letter +
                                   str(1)].value] = worksheet[column_letter + str(row)].value
        if row_data:
            excel_data.append(row_data)

    # Convert to JSON
    data = json.dumps(excel_data, sort_keys=True, indent=4)
    logging.info(f"Converted to JSON: {data}")
