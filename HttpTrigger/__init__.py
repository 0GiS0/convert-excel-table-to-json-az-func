import logging

import azure.functions as func

from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import json


def main(req: func.HttpRequest) -> func.HttpResponse:
    logging.info('Python HTTP trigger function processed a request.')

    for input_file in req.files.values():
        filename = input_file.filename

        logging.info('Processing: %s' % filename)

        # Load the workbook
        workbook = load_workbook(filename=BytesIO(input_file.stream.read()))
        logging.info(f"Loaded workbook: {workbook}")

        # Get the active sheet
        worksheet = workbook.active

        excel_data = []

        # Iterate through the rows and columns of the worksheet
        last_column = len(list(worksheet.columns))        
        logging.info(f"Last column: {last_column}")        
        last_row = len(list(worksheet.rows))
        logging.info(f"Last row: {last_row}")
                

        # Iterate row
        for row in range(1, last_row + 1):
            row_data = {}
            # Iterate column
            for column in range(1, last_column + 1):
                # Get the value of the cell
                column_letter = get_column_letter(column)
                logging.info(f"Column letter: {column_letter}")
                if row > 1:
                    logging.info(f"{worksheet[column_letter + str(1)].value}")
                    row_data[worksheet[column_letter + str(1)].value] = worksheet[column_letter + str(row)].value

            if row_data:
                excel_data.append(row_data)

        # Convert to JSON
        data = json.dumps(excel_data, sort_keys=True, indent=4)

    return func.HttpResponse(f'{data}')
